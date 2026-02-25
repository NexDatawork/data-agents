import numpy as np
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


def _synthetic_fill_numeric(
    df: pd.DataFrame,
    col: str,
    seed: int = 42,
    method: str = "lognormal",          
    clip_quantiles: tuple = (0.01, 0.99),
    expand: tuple = (0.5, 2.0),
    nonnegative: bool = True,
    integer: bool = True,
):
    """
    Fill ONLY missing values in `col` using synthetic data.
    Returns: (out_df, filled_mask, audit_dict)
    """
    rng = np.random.default_rng(seed)
    out = df.copy()
    
    # Protection, in case wrong column
    if col not in out.columns:
        raise KeyError(f"Column '{col}' not found in df.")

    # turning to numerical value
    # Look for na(NULL)
    s = pd.to_numeric(out[col], errors="coerce")
    na_mask = s.isna()

    if not na_mask.any():
        return out, na_mask, {"filled": 0, "method": method, "note": "no missing"}

    #  observed is the sample w/o missing value
    observed = s.dropna()
    if nonnegative:
        observed = observed[observed >= 0]

    # fallback if too few observed values
    # Use the lognormal to approximate a reasonable value
    if len(observed) < 10:
        if col.lower().startswith("inc"):
            base = rng.lognormal(mean=np.log(60000), sigma=0.7, size=int(na_mask.sum()))
        else:
            base = rng.lognormal(mean=np.log(8000), sigma=0.9, size=int(na_mask.sum()))
        gen = base
        bounds = {"low": None, "high": None}
        used = "fallback_lognormal"
    else:
        # based on existing data, 
        # select a range (strict or loose) to avoid outliers + align the trends
        q_low, q_high = observed.quantile([clip_quantiles[0], clip_quantiles[1]])
        low = max(q_low * expand[0], 0.0 if nonnegative else -np.inf)
        high = q_high * expand[1]
        bounds = {"low": float(low), "high": float(high)}

        n = int(na_mask.sum())

        # random generation
        if method == "empirical":
            # more conservative method
            gen = rng.choice(observed.to_numpy(), size=n, replace=True)
            used = "empirical"
        else:
            obs_pos = observed[observed > 0]
            # another protection on small sample size
            if len(obs_pos) < 10:
                gen = rng.choice(observed.to_numpy(), size=n, replace=True)
                used = "empirical_fallback"
            else:
                # generate lognormal distribution
                # i.e take log and calculate the mean, sd, then generate
                logx = np.log(obs_pos.to_numpy())
                mu = logx.mean()
                sigma = max(logx.std(ddof=1), 1e-6)
                gen = rng.lognormal(mean=mu, sigma=sigma, size=n)
                used = "lognormal"

        # avoid generating outliers envenmore
        gen = np.clip(gen, low, high)

    # make sure it's an integer, aligning with the real data
    if integer:
        gen = np.rint(gen).astype(int)
        if nonnegative:
            gen = np.maximum(gen, 0)

    out.loc[na_mask, col] = gen

    #  know what were filled and how
    audit = {
        "column": col,
        "filled": int(na_mask.sum()),
        "method": used,
        "bounds": bounds,
        "observed_n": int(len(observed)),
        "integer": integer,
    }
    return out, na_mask, audit

   
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
from typing import cast


def save_xlsx_with_highlights(
    df: pd.DataFrame,
    xlsx_path: str,
    highlight_map: dict,
    fill_color: str = "FFFF00",
):
    wb = Workbook()
    ws = cast(Worksheet, wb.active)   

    ws.title = "data"

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    yellow = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    col_to_excel = {name: i + 1 for i, name in enumerate(df.columns)}
    nrows = len(df)

    for col_name, mask in highlight_map.items():
        if col_name not in col_to_excel:
            continue
        c = col_to_excel[col_name]
        mask = pd.Series(mask).fillna(False).to_numpy()

        for i in range(nrows):
            if mask[i]:
                ws.cell(row=i + 2, column=c).fill = yellow

    wb.save(xlsx_path)



def main():
    # ---- change paths ----
    path_in = r"Your-csv-file-path-here.csv"
    path_out_xlsx = r"Your-output-xlsx-file-path-here.xlsx"
    path_out_csv = r"Your-output-csv-file-path-here.csv"

    df = pd.read_csv(path_in)

    # 1) Fill Income
    df1, income_mask, income_audit = _synthetic_fill_numeric(
        df, col="Income", seed=42, method="lognormal", integer=True, nonnegative=True
    )
    print(income_audit)

    # 2) Fill Loan_Balance
    df2, loan_mask, loan_audit = _synthetic_fill_numeric(
        df1, col="Loan_Balance", seed=42, method="lognormal", integer=True, nonnegative=True
    )
    print(loan_audit)
#
    df2["Income"] = pd.to_numeric(df2["Income"], errors="coerce").round().astype("Int64")
    df2["Loan_Balance"] = pd.to_numeric(df2["Loan_Balance"], errors="coerce").round().astype("Int64")

    # 3) Save CSV (no highlighting, but integer preserved)
    df2.to_csv(path_out_csv, index=False)
    print("Saved CSV:", path_out_csv)

    # 4) Save XLSX with yellow highlights for synthetic cells
    highlight_map = {
        "Income": income_mask,
        "Loan_Balance": loan_mask,
    }
    save_xlsx_with_highlights(df2, path_out_xlsx, highlight_map)
    print("Saved XLSX (highlighted):", path_out_xlsx)


if __name__ == "__main__":
    main()
